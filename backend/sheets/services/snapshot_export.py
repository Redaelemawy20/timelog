from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_snapshot_export(snapshot: dict) -> tuple[BytesIO, str]:
    sprints = snapshot.get("sprints", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sprints"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    separator_fill = PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="center")

    ws.append(["Start Date", "End Date", "Summary", "Time (hr)", "Projects"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 2
    for idx, sprint in enumerate(sprints):
        from_date = datetime.strptime(sprint["range_start"], "%Y-%m-%d").strftime("%d %b")
        to_date = datetime.strptime(sprint["range_end"], "%Y-%m-%d").strftime("%d %b")
        hours = float(sprint["time_hours"]) if sprint.get("time_hours") else 0
        projects = ", ".join(sprint.get("projects") or [])

        ws.append([from_date, to_date, sprint.get("summary", ""), hours, projects])

        for col_idx, cell in enumerate(ws[current_row], start=1):
            if col_idx == 3:
                cell.alignment = wrap_alignment
            else:
                cell.alignment = center_alignment

        if idx < len(sprints) - 1:
            current_row += 1
            ws.append(["", "", "", "", ""])
            for cell in ws[current_row]:
                cell.fill = separator_fill
            ws.row_dimensions[current_row].height = 8

        current_row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 150
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 25

    for row in ws.iter_rows(min_row=2, max_row=current_row):
        if row[0].value:
            ws.row_dimensions[row[0].row].height = None

    summary_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    summary_font = Font(bold=True)

    first_data_row = 2
    last_data_row = current_row
    sum_formula = f"=SUM(D{first_data_row}:D{last_data_row})"

    current_row += 1
    ws.append(["", "", "", sum_formula, ""])
    ws[f"D{current_row}"].fill = summary_fill
    ws[f"D{current_row}"].font = summary_font
    ws[f"D{current_row}"].alignment = center_alignment
    ws[f"C{current_row}"] = "Sum"
    ws[f"C{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"C{current_row}"].font = summary_font

    current_row += 1
    ws.append(["", "", "", 0, ""])
    ws[f"D{current_row}"].fill = summary_fill
    ws[f"D{current_row}"].font = summary_font
    ws[f"D{current_row}"].alignment = center_alignment
    ws[f"C{current_row}"] = "Previous"
    ws[f"C{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"C{current_row}"].font = summary_font

    current_row += 1
    sum_row = current_row - 2
    previous_row = current_row - 1
    total_formula = f"=D{sum_row}+D{previous_row}"
    ws.append(["", "", "", total_formula, ""])
    ws[f"D{current_row}"].fill = summary_fill
    ws[f"D{current_row}"].font = summary_font
    ws[f"D{current_row}"].alignment = center_alignment
    ws[f"C{current_row}"] = "Total"
    ws[f"C{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"C{current_row}"].font = summary_font

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    sheet_name_safe = snapshot.get("sheet_name", "sheet").replace(" ", "_").replace("/", "_").replace("\\", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{sheet_name_safe}_sprints_{timestamp}.xlsx"
    return buffer, filename
