from datetime import datetime
from io import BytesIO

from django.conf import settings
from django.db import transaction
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.request import Request
from rest_framework.response import Response

from .github_token import (
    fetch_commits_for_range,
    fetch_github_user_repos,
    read_github_token,
    resolve_token_path,
    verify_github_token,
)
from .llm_summary import (
    INITIAL_USER_MESSAGE,
    SummaryGenerationConfigError,
    SummaryGenerationUpstreamError,
    generate_conversation_reply,
)
from .models import Sheet, SheetRepo, Sprint, SprintConversationMessage, SprintRepo
from .serializers import (
    SheetCreateSerializer,
    SheetDetailSerializer,
    SheetListSerializer,
    SprintConversationMessageSerializer,
    SprintConversationSendSerializer,
    SprintCreateSerializer,
    SprintUpdateSerializer,
    SprintSerializer,
)


@api_view(["GET", "POST"])
def api_sheet_list(request: Request) -> Response:
    if request.method == "GET":
        sheets = Sheet.objects.all()
        return Response(SheetListSerializer(sheets, many=True).data)

    create = SheetCreateSerializer(data=request.data)
    create.is_valid(raise_exception=True)
    sheet = create.save()
    return Response(SheetListSerializer(sheet).data, status=status.HTTP_201_CREATED)


@api_view(["GET"])
def api_sheet_detail(request: Request, sheet_id: int) -> Response:
    sheet = get_object_or_404(Sheet, pk=sheet_id)
    return Response(SheetDetailSerializer(sheet).data)


@api_view(["POST"])
def api_sprint_create(request: Request, sheet_id: int) -> Response:
    sheet = get_object_or_404(Sheet, pk=sheet_id)
    create = SprintCreateSerializer(data=request.data)
    create.is_valid(raise_exception=True)
    data = create.validated_data
    range_start = data["range_start"]
    range_end = data["range_end"]
    repo_specs = data["repos"]

    resolved_path = resolve_token_path(settings.BASE_DIR, settings.GITHUB_TOKEN_FILE)
    if resolved_path is None:
        return Response(
            {"detail": "GitHub token file is not configured (GITHUB_TOKEN_FILE)."},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )
    read = read_github_token(resolved_path)
    if not read.ok:
        return Response(
            {"detail": read.error},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )
    assert read.token is not None

    prepared: list[tuple[str, str, str, str, list[dict]]] = []
    for spec in repo_specs:
        owner = spec["owner"]
        name = spec["name"]
        display_name_raw = spec.get("display_name")
        display_name = ""
        if isinstance(display_name_raw, str):
            display_name = display_name_raw.strip()
        if not display_name:
            display_name = f"{owner}/{name}"
        branch_raw = spec.get("default_branch")
        branch = "main"
        if isinstance(branch_raw, str) and branch_raw.strip():
            branch = branch_raw.strip()

        listed = fetch_commits_for_range(read.token, owner, name, branch, range_start, range_end)
        if not listed.ok:
            gh_code = listed.http_status
            if gh_code == 401:
                resp_status = status.HTTP_401_UNAUTHORIZED
            elif gh_code == 403:
                resp_status = status.HTTP_403_FORBIDDEN
            elif gh_code is not None and 500 <= gh_code < 600:
                resp_status = status.HTTP_502_BAD_GATEWAY
            else:
                resp_status = status.HTTP_502_BAD_GATEWAY
            return Response({"detail": listed.error}, status=resp_status)

        commits = listed.commits or []
        prepared.append((owner, name, display_name, branch, commits))

    with transaction.atomic():
        sprint = Sprint.objects.create(
            sheet=sheet,
            range_start=range_start,
            range_end=range_end,
            status=Sprint.Status.DRAFT,
            summary="",
            time_hours=None,
        )
        for owner, name, display_name, _branch, commits in prepared:
            sheet_repo, _ = SheetRepo.objects.get_or_create(
                sheet=sheet,
                owner=owner,
                name=name,
                defaults={"display_name": display_name},
            )
            messages_joined = "\n\n".join(
                c["message"] for c in commits if isinstance(c.get("message"), str) and c["message"]
            )
            SprintRepo.objects.create(
                sheet=sheet,
                sprint=sprint,
                sheet_repo=sheet_repo,
                project=name,
                notes=None,
                commit_messages=messages_joined,
                raw_commits_json=commits,
            )

    sprint = Sprint.objects.prefetch_related("sprint_repos__sheet_repo").get(pk=sprint.pk)
    return Response(SprintSerializer(sprint).data, status=status.HTTP_201_CREATED)


@api_view(["PATCH", "DELETE"])
def api_sprint_update(request: Request, sheet_id: int, sprint_id: int) -> Response:
    sprint = get_object_or_404(Sprint, pk=sprint_id, sheet_id=sheet_id)
    
    if request.method == "DELETE":
        sprint.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    update = SprintUpdateSerializer(data=request.data)
    update.is_valid(raise_exception=True)
    changed_fields: list[str] = []
    if "summary" in update.validated_data:
        sprint.summary = update.validated_data["summary"]
        changed_fields.append("summary")
    if "time_hours" in update.validated_data:
        sprint.time_hours = update.validated_data["time_hours"]
        changed_fields.append("time_hours")
    sprint.save(update_fields=changed_fields)
    return Response(SprintSerializer(sprint).data)


def _sprint_with_repos(sprint_id: int) -> Sprint:
    return get_object_or_404(
        Sprint.objects.prefetch_related("sprint_repos__sheet_repo"),
        pk=sprint_id,
    )


def _conversation_messages(sprint: Sprint) -> list[SprintConversationMessage]:
    return list(sprint.conversation_messages.order_by("created_at"))


def _conversation_error_response(exc: Exception) -> Response:
    if isinstance(exc, SummaryGenerationConfigError):
        return Response({"detail": str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
    if isinstance(exc, SummaryGenerationUpstreamError):
        return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
    raise exc


@api_view(["GET", "POST"])
def api_sprint_conversation(request: Request, sprint_id: int) -> Response:
    sprint = _sprint_with_repos(sprint_id)

    if request.method == "GET":
        messages = _conversation_messages(sprint)
        return Response(SprintConversationMessageSerializer(messages, many=True).data)

    send = SprintConversationSendSerializer(data=request.data)
    send.is_valid(raise_exception=True)
    content = send.validated_data["content"]
    init = send.validated_data["init"]
    existing = _conversation_messages(sprint)

    if init or (not existing and not content):
        user_content = INITIAL_USER_MESSAGE
    elif content:
        user_content = content
    else:
        return Response(
            {"detail": "Message content is required."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user_message = SprintConversationMessage.objects.create(
        sprint=sprint,
        role=SprintConversationMessage.Role.USER,
        content=user_content,
    )
    history = existing + [user_message]

    try:
        reply = generate_conversation_reply(sprint, history)
    except (SummaryGenerationConfigError, SummaryGenerationUpstreamError) as exc:
        user_message.delete()
        return _conversation_error_response(exc)

    assistant_message = SprintConversationMessage.objects.create(
        sprint=sprint,
        role=SprintConversationMessage.Role.ASSISTANT,
        content=reply,
    )
    return Response(
        SprintConversationMessageSerializer([user_message, assistant_message], many=True).data,
        status=status.HTTP_201_CREATED,
    )


@api_view(["GET"])
def api_github_token_status(request: Request) -> Response:
    resolved_path = resolve_token_path(settings.BASE_DIR, settings.GITHUB_TOKEN_FILE)
    if resolved_path is None:
        return Response({"phase": "not_configured"})

    read = read_github_token(resolved_path)
    if not read.ok:
        return Response({"phase": "file_error", "file_error": read.error})

    assert read.token is not None
    verified = verify_github_token(read.token)
    if verified.ok:
        return Response({"phase": "success", "login": verified.login, "name": verified.name})

    return Response({"phase": "api_error", "api_error": verified.error})


@api_view(["GET"])
def api_github_repos(request: Request) -> Response:
    """List GitHub repositories for the configured user token (GITHUB_TOKEN_FILE)."""
    resolved_path = resolve_token_path(settings.BASE_DIR, settings.GITHUB_TOKEN_FILE)
    if resolved_path is None:
        return Response(
            {"error": "GitHub token file is not configured (GITHUB_TOKEN_FILE)."},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    read = read_github_token(resolved_path)
    if not read.ok:
        return Response(
            {"error": read.error},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    assert read.token is not None
    listed = fetch_github_user_repos(read.token)
    if not listed.ok:
        gh_code = listed.http_status
        if gh_code == 401:
            resp_status = status.HTTP_401_UNAUTHORIZED
        elif gh_code == 403:
            resp_status = status.HTTP_403_FORBIDDEN
        elif gh_code is not None and 500 <= gh_code < 600:
            resp_status = status.HTTP_502_BAD_GATEWAY
        else:
            resp_status = status.HTTP_502_BAD_GATEWAY
        return Response({"error": listed.error}, status=resp_status)

    assert listed.repos is not None
    return Response({"repos": listed.repos})


@api_view(["GET"])
def api_sheet_export_excel(request: Request, sheet_id: int) -> FileResponse:
    sheet = get_object_or_404(Sheet, pk=sheet_id)
    sprints = sheet.sprints.all().prefetch_related("sprint_repos__sheet_repo").order_by("range_start")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sprints"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    separator_fill = PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="center")

    ws.append(["Start Date", "End Date", "Summary", "Time (hr)", "Projects"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 2
    for idx, sprint in enumerate(sprints):
        projects = [
            repo.project
            for repo in sprint.sprint_repos.all()
            if repo.commit_messages and repo.commit_messages.strip()
        ]

        ws.append([
            sprint.range_start.strftime("%d %b"),
            sprint.range_end.strftime("%d %b"),
            sprint.summary,
            float(sprint.time_hours) if sprint.time_hours else 0,
            ", ".join(projects),
        ])

        for col_idx, cell in enumerate(ws[current_row], start=1):
            if col_idx == 3:
                cell.alignment = wrap_alignment
            else:
                cell.alignment = center_alignment

        if idx < len(sprints) - 1:
            current_row += 1
            ws.append(["", "", "", "", ""])
            for cell in ws[current_row]:
                cell.fill = separator_fill
            ws.row_dimensions[current_row].height = 8

        current_row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 150
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 25

    for row in ws.iter_rows(min_row=2, max_row=current_row):
        if row[0].value:
            ws.row_dimensions[row[0].row].height = None

    summary_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    summary_font = Font(bold=True)
    
    first_data_row = 2
    last_data_row = current_row
    sum_formula = f"=SUM(D{first_data_row}:D{last_data_row})"
    
    current_row += 1
    ws.append(["", "", "", sum_formula, ""])
    ws[f"D{current_row}"].fill = summary_fill
    ws[f"D{current_row}"].font = summary_font
    ws[f"D{current_row}"].alignment = center_alignment
    ws[f"C{current_row}"] = "Sum"
    ws[f"C{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"C{current_row}"].font = summary_font
    
    current_row += 1
    ws.append(["", "", "", 0, ""])
    ws[f"D{current_row}"].fill = summary_fill
    ws[f"D{current_row}"].font = summary_font
    ws[f"D{current_row}"].alignment = center_alignment
    ws[f"C{current_row}"] = "Previous"
    ws[f"C{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"C{current_row}"].font = summary_font
    
    current_row += 1
    sum_row = current_row - 2
    previous_row = current_row - 1
    total_formula = f"=D{sum_row}+D{previous_row}"
    ws.append(["", "", "", total_formula, ""])
    ws[f"D{current_row}"].fill = summary_fill
    ws[f"D{current_row}"].font = summary_font
    ws[f"D{current_row}"].alignment = center_alignment
    ws[f"C{current_row}"] = "Total"
    ws[f"C{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"C{current_row}"].font = summary_font

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    sheet_name_safe = sheet.name.replace(" ", "_").replace("/", "_").replace("\\", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{sheet_name_safe}_sprints_{timestamp}.xlsx"

    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
